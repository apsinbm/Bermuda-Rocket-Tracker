#!/usr/bin/env python3
"""
Hybrid E-commerce Monitor
Uses web scraping with 20-second delays for Freshop stores (MP, HH, Miles, Drop It)
Uses API for Pronto store
Includes OCR fallback capability
"""

import pandas as pd
import time
from datetime import datetime
import logging
import random
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import undetected_chromedriver as uc
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import sys
import os
from api_integrations import EddressAPI
import easyocr
import cv2
import numpy as np
from PIL import Image
import io

class HybridEcommerceMonitor:
    def __init__(self, input_file, output_file="Hybrid_Monitor_Results.xlsx"):
        self.input_file = input_file
        self.output_file = output_file
        self.results = []
        self.auto_save_counter = 0
        self.auto_save_interval = 10  # Save every 10 results
        
        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('hybrid_monitor.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
        # Initialize Pronto API
        self.pronto_api = EddressAPI()
        
        # Initialize OCR reader
        try:
            self.ocr_reader = easyocr.Reader(['en'])
            self.ocr_available = True
            self.logger.info("OCR system initialized successfully")
        except Exception as e:
            self.logger.warning(f"OCR initialization failed: {e}")
            self.ocr_available = False
        
        # Chrome options for stealth browsing
        self.viewport_sizes = [
            (1920, 1080), (1366, 768), (1536, 864),
            (1440, 900), (1280, 720), (1600, 900)
        ]
        
        self.user_agents = [
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Safari/605.1.15",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/119.0.0.0"
        ]
        
        self.performance_stats = {
            'total_requests': 0,
            'successful_scrapes': 0,
            'successful_apis': 0,
            'successful_ocr': 0,
            'failed_requests': 0
        }
    
    def auto_save_results(self):
        """Auto-save results every 10 processed items"""
        try:
            if len(self.results) > 0:
                # Create auto-save filename with timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                auto_save_file = f"AutoSave_{timestamp}_{len(self.results)}_results.xlsx"
                
                # Generate report with current results
                self.generate_report(auto_save_file)
                self.logger.info(f"🔄 Auto-saved {len(self.results)} results to {auto_save_file}")
        except Exception as e:
            self.logger.warning(f"Auto-save failed: {e}")
    
    def add_result_with_autosave(self, result):
        """Add result and trigger auto-save if needed"""
        self.results.append(result)
        self.auto_save_counter += 1
        
        # Auto-save every 10 results
        if self.auto_save_counter >= self.auto_save_interval:
            self.auto_save_results()
            self.auto_save_counter = 0
    
    def create_driver(self):
        """Create a stealth Chrome driver with random configurations"""
        try:
            # Random configurations
            width, height = random.choice(self.viewport_sizes)
            user_agent = random.choice(self.user_agents)
            
            # Chrome options - avoid profile conflicts for now  
            options = Options()
            
            # Comment out profile for now to avoid conflicts
            # options.add_argument("--user-data-dir=/Users/pato/Library/Application Support/Google/Chrome")
            # options.add_argument("--profile-directory=Automation")
            
            options.add_argument(f'--user-agent={user_agent}')
            options.add_argument(f'--window-size={width},{height}')
            options.add_argument('--disable-web-security')
            options.add_argument('--disable-features=VizDisplayCompositor')
            options.add_argument('--disable-extensions')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            # Try undetected Chrome first
            try:
                driver = uc.Chrome(options=options, version_main=None)
                self.logger.info("Created undetected Chrome driver")
            except Exception as e:
                self.logger.warning(f"Undetected Chrome failed: {e}, using webdriver-manager Chrome")
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=options)
            
            # Remove webdriver property
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            return driver
            
        except Exception as e:
            self.logger.error(f"Failed to create driver: {e}")
            return None
    
    def random_mouse_movement(self, driver):
        """Simulate random mouse movements"""
        try:
            actions = ActionChains(driver)
            
            # Get viewport size
            viewport_width = driver.execute_script("return window.innerWidth")
            viewport_height = driver.execute_script("return window.innerHeight")
            
            # Random mouse movements
            for _ in range(random.randint(2, 5)):
                x = random.randint(0, viewport_width)
                y = random.randint(0, viewport_height)
                actions.move_by_offset(x, y)
                time.sleep(random.uniform(0.1, 0.3))
            
            actions.perform()
            
        except Exception as e:
            self.logger.debug(f"Mouse movement failed: {e}")
    
    def random_scroll(self, driver):
        """Simulate random scrolling"""
        try:
            # Random scroll
            scroll_amount = random.randint(300, 800)
            if random.choice([True, False]):
                scroll_amount = -scroll_amount
            
            driver.execute_script(f"window.scrollBy(0, {scroll_amount});")
            time.sleep(random.uniform(0.5, 1.5))
            
        except Exception as e:
            self.logger.debug(f"Scrolling failed: {e}")
    
    def extract_comprehensive_pricing(self, driver):
        """Extract comprehensive pricing information including sale prices, regular prices, and sale dates"""
        pricing_info = {
            'current_price': None,
            'sale_price': None,
            'regular_price': None,
            'original_price': None,
            'sale_percentage': None,
            'sale_start_date': None,
            'sale_end_date': None,
            'is_on_sale': False,
            'price_method': 'none',
            'website_sku': None
        }
        
        try:
            # First check if this is MarketPlace - they have specific CSS classes
            current_url = driver.current_url
            if 'marketplace.bm' in current_url:
                return self._extract_marketplace_pricing(driver)
            else:
                return self._extract_generic_pricing(driver)
        except Exception as e:
            self.logger.debug(f"Comprehensive pricing extraction failed: {e}")
            return pricing_info
    
    def _extract_marketplace_pricing(self, driver):
        """Extract pricing specifically for MarketPlace using their CSS classes"""
        pricing_info = {
            'current_price': None,
            'sale_price': None,
            'regular_price': None,
            'original_price': None,
            'sale_percentage': None,
            'sale_start_date': None,
            'sale_end_date': None,
            'is_on_sale': False,
            'price_method': 'marketplace_specific',
            'website_sku': None
        }
        
        try:
            # Step 1: Look for regular/base price first (crossed out)
            regular_price = None
            regular_elements = driver.find_elements(By.CSS_SELECTOR, '.fp-item-base-price')
            self.logger.info(f"🔍 Found {len(regular_elements)} base price elements")
            for elem in regular_elements:
                text = elem.text.strip()
                self.logger.info(f"   Base price text: '{text}'")
                match = re.search(r'\$(\d+\.?\d*)', text)
                if match:
                    price = float(match.group(1))
                    self.logger.info(f"   Extracted price: ${price}")
                    # Check if this element has strikethrough
                    try:
                        computed_style = driver.execute_script(
                            "return window.getComputedStyle(arguments[0]).textDecoration;", elem
                        )
                        if 'line-through' in (computed_style or ''):
                            regular_price = price
                            pricing_info['regular_price'] = price
                            pricing_info['original_price'] = price
                            self.logger.info(f"✓ MarketPlace regular price found (strikethrough): ${price}")
                            break
                    except:
                        pass
                    
                    # For regular products (no sale), use first valid price as regular price
                    if not regular_price and 0.01 <= price <= 1000:
                        regular_price = price
                        pricing_info['regular_price'] = price
                        pricing_info['original_price'] = price
                        self.logger.info(f"✓ MarketPlace regular price found: ${price}")
                        break
            
            # Step 2: Look for sale price using MarketPlace-specific classes  
            sale_price = None
            # Try multiple selectors for sale price
            sale_selectors = [
                '.fp-item-sale strong',  # <strong>$8.99</strong>
                '.fp-item-sale b',       # <b>$8.99</b>
                '.fp-item-sale span',    # <span>$8.99</span>
                '.fp-item-sale'          # Direct in sale element
            ]
            
            for selector in sale_selectors:
                sale_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in sale_elements:
                    text = elem.text.strip()
                    match = re.search(r'\$(\d+\.?\d*)', text)
                    if match:
                        price = float(match.group(1))
                        # Verify this is different from regular price and reasonable
                        if regular_price and price < regular_price and 0.01 <= price <= 1000:
                            sale_price = price
                            pricing_info['sale_price'] = price
                            pricing_info['current_price'] = price
                            pricing_info['is_on_sale'] = True
                            self.logger.info(f"✓ MarketPlace sale price found: ${price}")
                            break
                        elif not regular_price and 0.01 <= price <= 1000:
                            # No regular price found yet, use this as current price
                            sale_price = price
                            pricing_info['current_price'] = price
                if sale_price:
                    break
            
            # Step 3: If no explicit sale price found but we have regular price, look for prominent price
            if not sale_price and regular_price:
                # Look for other prominent prices that might be the sale price
                prominent_selectors = ['h1', 'h2', 'h3', '.price', '[class*="price"]']
                for selector in prominent_selectors:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for elem in elements:
                        text = elem.text.strip()
                        match = re.search(r'\$(\d+\.?\d*)', text)
                        if match:
                            price = float(match.group(1))
                            if price < regular_price and 0.01 <= price <= 1000:
                                sale_price = price
                                pricing_info['sale_price'] = price
                                pricing_info['current_price'] = price
                                pricing_info['is_on_sale'] = True
                                self.logger.info(f"✓ MarketPlace sale price found in {selector}: ${price}")
                                break
                    if sale_price:
                        break
                        
            # Extract sale dates from MarketPlace format (06/25/25 - 07/06/25)
            sale_date_elements = driver.find_elements(By.CSS_SELECTOR, '.fp-item-sale-date, .fp-item-sale')
            for elem in sale_date_elements:
                text = elem.text
                date_match = re.search(r'\((\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})\)', text)
                if date_match:
                    pricing_info['sale_start_date'] = date_match.group(1)
                    pricing_info['sale_end_date'] = date_match.group(2)
                    break
                    
            # CRITICAL FIX: If we have regular price but no sale price, use regular price as current price
            if pricing_info['regular_price'] and not pricing_info['current_price']:
                pricing_info['current_price'] = pricing_info['regular_price']
                self.logger.info(f"✓ MarketPlace current price: ${pricing_info['regular_price']} (no sale)")
            
            # Calculate sale percentage if we have both prices
            if pricing_info['sale_price'] and pricing_info['regular_price']:
                sale_pct = ((pricing_info['regular_price'] - pricing_info['sale_price']) / pricing_info['regular_price']) * 100
                pricing_info['sale_percentage'] = round(sale_pct, 1)
                pricing_info['is_on_sale'] = True
                self.logger.info(f"✓ MarketPlace sale: ${pricing_info['sale_price']} (was ${pricing_info['regular_price']}) = {pricing_info['sale_percentage']}% OFF")
            
            # Extract SKU/UPC from MarketPlace page
            try:
                sku_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'SKU') or contains(text(), 'UPC') or contains(text(), 'Item#') or contains(text(), 'Product Code')]")
                for elem in sku_elements:
                    text = elem.text
                    # Look for patterns like "SKU: 123456" or "UPC: 123456789" or "Item# 123456"
                    sku_patterns = [
                        r'SKU[:\s#]*(\d+)',
                        r'UPC[:\s#]*(\d+)',
                        r'Item[#\s]*(\d+)',
                        r'Product Code[:\s]*(\d+)',
                        r'PLU[:\s#]*(\d+)'
                    ]
                    for pattern in sku_patterns:
                        match = re.search(pattern, text, re.IGNORECASE)
                        if match:
                            pricing_info['website_sku'] = match.group(1)
                            self.logger.info(f"✓ MarketPlace SKU found: {match.group(1)}")
                            break
                    if pricing_info['website_sku']:
                        break
            except Exception:
                pass
                
            return pricing_info
            
        except Exception as e:
            self.logger.debug(f"MarketPlace pricing extraction failed: {e}")
            return pricing_info
    
    def _extract_generic_pricing(self, driver):
        """Extract pricing for non-MarketPlace stores using generic selectors"""
        pricing_info = {
            'current_price': None,
            'sale_price': None,
            'regular_price': None,
            'original_price': None,
            'sale_percentage': None,
            'sale_start_date': None,
            'sale_end_date': None,
            'is_on_sale': False,
            'price_method': 'generic',
            'website_sku': None
        }
        
        try:
            # Simplified generic pricing for stores like Pronto
            # Look for strikethrough prices (regular prices)
            strikethrough_elements = driver.execute_script("""
                var elements = Array.from(document.querySelectorAll('*'));
                return elements.filter(el => {
                    var style = window.getComputedStyle(el);
                    var text = el.textContent;
                    return (style.textDecoration.includes('line-through') || 
                           el.tagName === 'DEL' || el.tagName === 'S') && 
                           text.includes('$') && el.offsetParent !== null;
                }).map(el => ({
                    text: el.textContent.trim(),
                    tag: el.tagName
                }));
            """)
            
            # Extract regular price from strikethrough elements
            for elem in strikethrough_elements:
                text = elem['text']
                match = re.search(r'\$(\d+\.?\d*)', text)
                if match:
                    pricing_info['regular_price'] = float(match.group(1))
                    pricing_info['original_price'] = pricing_info['regular_price']
                    break
            
            # Look for sale percentage indicators (like "35% OFF")
            sale_elements = driver.execute_script("""
                var elements = Array.from(document.querySelectorAll('*'));
                return elements.filter(el => {
                    var text = el.textContent.toLowerCase();
                    return text.match(/\\d+%\\s*(off|sale)/) && el.offsetParent !== null;
                }).map(el => el.textContent.trim());
            """)
            
            for text in sale_elements:
                percentage_match = re.search(r'(\d+)%\s*(off|sale)', text, re.IGNORECASE)
                if percentage_match:
                    pricing_info['sale_percentage'] = int(percentage_match.group(1))
                    pricing_info['is_on_sale'] = True
                    break
            
            # Look for current price (largest, most prominent text)
            price_elements = driver.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, .price, [class*="price"]')
            best_price = None
            best_score = 0
            
            for elem in price_elements:
                try:
                    text = elem.text.strip()
                    match = re.search(r'\$(\d+\.?\d*)', text)
                    if match:
                        price = float(match.group(1))
                        if 0.01 <= price <= 1000:
                            # Score based on element prominence
                            score = 0
                            if elem.tag_name in ['h1', 'h2', 'h3']:
                                score += 10
                            
                            # Check font size and color via JavaScript
                            try:
                                element_info = driver.execute_script("""
                                    var style = window.getComputedStyle(arguments[0]);
                                    return {
                                        fontSize: parseInt(style.fontSize),
                                        color: style.color,
                                        fontWeight: style.fontWeight
                                    };
                                """, elem)
                                
                                if element_info['fontSize'] > 16:
                                    score += element_info['fontSize'] // 4
                                if 'bold' in str(element_info['fontWeight']) or int(element_info.get('fontWeight', 400)) >= 600:
                                    score += 5
                                if 'blue' in element_info.get('color', '').lower():
                                    score += 5
                            except:
                                pass
                            
                            if score > best_score:
                                best_score = score
                                best_price = price
                except:
                    continue
            
            if best_price:
                pricing_info['current_price'] = best_price
                if pricing_info['is_on_sale']:
                    pricing_info['sale_price'] = best_price
                    
            # Calculate sale percentage if we have both prices but not explicit percentage
            if (pricing_info['regular_price'] and pricing_info['current_price'] and 
                pricing_info['regular_price'] > pricing_info['current_price'] and 
                not pricing_info['sale_percentage']):
                sale_pct = ((pricing_info['regular_price'] - pricing_info['current_price']) / pricing_info['regular_price']) * 100
                pricing_info['sale_percentage'] = round(sale_pct, 1)
                pricing_info['is_on_sale'] = True
                pricing_info['sale_price'] = pricing_info['current_price']
                
            pricing_info['price_method'] = 'generic_scraping'
            
            # Extract SKU/UPC for generic stores (Pronto, HH, etc.)
            try:
                sku_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'SKU') or contains(text(), 'UPC') or contains(text(), 'Item#') or contains(text(), 'Product Code') or contains(text(), 'PLU')]")
                for elem in sku_elements:
                    text = elem.text
                    # Look for patterns like "SKU: 123456" or "UPC: 123456789" or "Item# 123456"
                    sku_patterns = [
                        r'SKU[:\s#]*(\d+)',
                        r'UPC[:\s#]*(\d+)',
                        r'Item[#\s]*(\d+)',
                        r'Product Code[:\s]*(\d+)',
                        r'PLU[:\s#]*(\d+)'
                    ]
                    for pattern in sku_patterns:
                        match = re.search(pattern, text, re.IGNORECASE)
                        if match:
                            pricing_info['website_sku'] = match.group(1)
                            self.logger.info(f"✓ SKU found: {match.group(1)}")
                            break
                    if pricing_info['website_sku']:
                        break
            except Exception:
                pass
                
        except Exception as e:
            self.logger.debug(f"Generic pricing extraction failed: {e}")
            pricing_info['price_method'] = 'extraction_failed'
        
        return pricing_info
    
    def select_hamilton_store(self, driver):
        """Select Hamilton store at the beginning of MarketPlace session"""
        try:
            # Step 1: Go to MarketPlace homepage
            self.logger.info("📍 Loading MarketPlace homepage...")
            driver.get("https://www.marketplace.bm")
            time.sleep(5)
            
            # Step 2: Look for and click "My Store"
            store_selectors = [
                "//a[contains(text(), 'My Store')]",
                "//button[contains(text(), 'My Store')]", 
                "//div[contains(text(), 'My Store')]"
            ]
            
            store_button = None
            for selector in store_selectors:
                try:
                    elements = driver.find_elements(By.XPATH, selector)
                    if elements:
                        store_button = elements[0]
                        self.logger.info(f"✓ Found store selector")
                        break
                except:
                    continue
            
            if store_button:
                self.logger.info("🔄 Clicking store selector...")
                store_button.click()
                time.sleep(3)
                
                # Step 3: Look for Hamilton store option and click it
                hamilton_selectors = [
                    "//a[contains(text(), 'Hamilton')]",
                    "//button[contains(text(), 'Hamilton')]",
                    "//div[contains(text(), 'Hamilton')]",
                    "//span[contains(text(), 'Hamilton')]"
                ]
                
                hamilton_found = False
                for selector in hamilton_selectors:
                    try:
                        elements = driver.find_elements(By.XPATH, selector)
                        if elements:
                            self.logger.info("✓ Found Hamilton store option")
                            elements[0].click()
                            time.sleep(3)
                            hamilton_found = True
                            break
                    except:
                        continue
                
                if hamilton_found:
                    self.logger.info("✅ Hamilton store selected successfully")
                    return True
                else:
                    self.logger.warning("⚠️  Hamilton store option not found")
                    return False
            else:
                self.logger.warning("⚠️  Store selector not found")
                return False
                
        except Exception as e:
            self.logger.error(f"❌ Store selection failed: {e}")
            return False
    
    def handle_store_selection(self, driver, store_type):
        """Handle store selection for MarketPlace and Drop It"""
        try:
            if store_type == 'mp':  # MarketPlace
                # Look for store selection elements
                store_selectors = [
                    "button:contains('Hamilton')",
                    "div:contains('The MarketPlace Hamilton')",
                    "div:contains('42 Church Street')",
                    "*[data-store*='hamilton']",
                    "*[class*='store-selector']",
                    "*[class*='location']",
                    ".store-location",
                    "#store-selector"
                ]
                
                for selector in store_selectors:
                    try:
                        if 'contains' in selector:
                            # Use JavaScript for text-based selection
                            element = driver.execute_script(f"""
                                var elements = Array.from(document.querySelectorAll('*'));
                                return elements.find(el => 
                                    el.textContent.includes('Hamilton') || 
                                    el.textContent.includes('42 Church Street') ||
                                    el.textContent.includes('The MarketPlace Hamilton')
                                );
                            """)
                            if element:
                                driver.execute_script("arguments[0].click();", element)
                                self.logger.info("Selected MarketPlace Hamilton store")
                                time.sleep(2)
                                break
                        else:
                            element = driver.find_element(By.CSS_SELECTOR, selector)
                            element.click()
                            self.logger.info("Selected MarketPlace Hamilton store")
                            time.sleep(2)
                            break
                    except:
                        continue
                        
            elif store_type == 'dropit':  # Drop It
                # Look for store selection elements
                store_selectors = [
                    "button:contains('Warwick')",
                    "div:contains('Warwick')",
                    "*[data-store*='warwick']",
                    "*[class*='store-selector']",
                    "*[class*='location']",
                    ".store-location",
                    "#store-selector"
                ]
                
                for selector in store_selectors:
                    try:
                        if 'contains' in selector:
                            # Use JavaScript for text-based selection
                            element = driver.execute_script(f"""
                                var elements = Array.from(document.querySelectorAll('*'));
                                return elements.find(el => 
                                    el.textContent.includes('Warwick')
                                );
                            """)
                            if element:
                                driver.execute_script("arguments[0].click();", element)
                                self.logger.info("Selected Drop It Warwick store")
                                time.sleep(2)
                                break
                        else:
                            element = driver.find_element(By.CSS_SELECTOR, selector)
                            element.click()
                            self.logger.info("Selected Drop It Warwick store")
                            time.sleep(2)
                            break
                    except:
                        continue
                        
        except Exception as e:
            self.logger.debug(f"Store selection handling failed for {store_type}: {e}")
    
    def extract_price_with_ocr(self, driver):
        """Extract price using OCR from screenshot"""
        if not self.ocr_available:
            return None
        
        try:
            # Take screenshot
            screenshot = driver.get_screenshot_as_png()
            image = Image.open(io.BytesIO(screenshot))
            
            # Convert to numpy array for OpenCV
            opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            
            # Extract text using OCR
            results = self.ocr_reader.readtext(opencv_image)
            
            # Look for price patterns
            price_patterns = [
                r'\$\s*(\d+\.?\d*)',  # $12.34 or $12
                r'(\d+\.?\d*)\s*\$',  # 12.34$ or 12$
                r'Price[:\s]*\$?\s*(\d+\.?\d*)',  # Price: $12.34
                r'USD[:\s]*\$?\s*(\d+\.?\d*)',    # USD: $12.34
            ]
            
            for detection in results:
                text = detection[1].strip()
                
                for pattern in price_patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        try:
                            price = float(match.group(1))
                            if 0.01 <= price <= 1000:  # Reasonable price range
                                self.logger.info(f"OCR found price: ${price}")
                                return price
                        except ValueError:
                            continue
            
            return None
            
        except Exception as e:
            self.logger.error(f"OCR extraction failed: {e}")
            return None
    
    def scrape_freshop_store(self, url, store_type, product_name, driver=None):
        """Scrape Freshop stores with 20-second delays and OCR fallback"""
        driver_created_here = False
        try:
            # Use provided driver or create new one
            if not driver:
                driver = self.create_driver()
                driver_created_here = True
                if not driver:
                    return {'error': 'Failed to create driver', 'method': 'scraping'}
            
            self.logger.info(f"🔍 Scraping {store_type.upper()}: {product_name}")
            
            # Navigate to URL
            driver.get(url)
            
            # Wait and simulate human behavior
            time.sleep(random.uniform(3, 5))
            self.random_mouse_movement(driver)
            time.sleep(random.uniform(2, 3))
            self.random_scroll(driver)
            
            # Wait for page to fully load
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            
            # Handle store selection for MarketPlace and Drop It
            self.handle_store_selection(driver, store_type)
            
            # Long delay as requested (20 seconds)
            self.logger.info(f"Waiting 20 seconds for {store_type.upper()} dynamic content...")
            time.sleep(20)
            
            # Use comprehensive pricing extraction to detect sales, regular prices, and dates
            pricing_info = self.extract_comprehensive_pricing(driver)
            price_found = pricing_info.get('current_price')
            method_used = 'comprehensive_scraping' if price_found else 'scraping'
            
            # If scraping failed, try OCR
            if not price_found and self.ocr_available:
                self.logger.info(f"Scraping failed for {store_type.upper()}, trying OCR...")
                price_found = self.extract_price_with_ocr(driver)
                if price_found:
                    method_used = 'ocr'
                    self.performance_stats['successful_ocr'] += 1
            
            if price_found:
                if method_used in ['scraping', 'comprehensive_scraping']:
                    self.performance_stats['successful_scrapes'] += 1
                
                # Return comprehensive pricing information
                result = {
                    'price_found': price_found,
                    'method': method_used,
                    'status': 'success'
                }
                
                # Add comprehensive pricing data if available
                if method_used == 'comprehensive_scraping':
                    result.update({
                        'sale_price': pricing_info.get('sale_price'),
                        'regular_price': pricing_info.get('regular_price'),
                        'original_price': pricing_info.get('original_price'),
                        'sale_percentage': pricing_info.get('sale_percentage'),
                        'sale_start_date': pricing_info.get('sale_start_date'),
                        'sale_end_date': pricing_info.get('sale_end_date'),
                        'is_on_sale': pricing_info.get('is_on_sale', False),
                        'website_sku': pricing_info.get('website_sku')
                    })
                
                return result
            else:
                self.logger.warning(f"❌ No price found for {store_type.upper()}: {product_name}")
                return {'error': 'No price found', 'method': 'scraping+ocr'}
                
        except Exception as e:
            self.logger.error(f"Error scraping {url}: {e}")
            return {'error': str(e), 'method': 'scraping'}
        
        finally:
            # Only close driver if we created it here (not for persistent MarketPlace sessions)
            if driver and driver_created_here:
                try:
                    driver.quit()
                except:
                    pass
    
    def get_pronto_data(self, url, product_name):
        """Get Pronto data using API (which works well)"""
        try:
            self.logger.info(f"🔗 API call for PRONTO: {product_name}")
            result = self.pronto_api.get_product_data(url)
            
            if result.get('price_found'):
                self.performance_stats['successful_apis'] += 1
                return {
                    'price_found': result['price_found'],
                    'method': 'api',
                    'status': 'success'
                }
            else:
                return {'error': result.get('error', 'No price found'), 'method': 'api'}
                
        except Exception as e:
            self.logger.error(f"Error with Pronto API for {url}: {e}")
            return {'error': str(e), 'method': 'api'}
    
    def process_single_url(self, url, store_type, product_info, product_name, driver=None):
        """Process a single URL using appropriate method"""
        if not url or pd.isna(url):
            return {
                'upc_plu': product_info.get('UPC/PLU', ''),
                'brand': product_info.get('Brand', ''),
                'product_name': product_name,
                'url': None,
                'store_type': store_type,
                'status': 'skipped',
                'error': 'No URL provided',
                'price_found': None,
                'expected_price': product_info.get('Reg Retail'),
                'method': 'none'
            }
        
        start_time = time.time()
        self.performance_stats['total_requests'] += 1
        
        try:
            # Use comprehensive web scraping for all stores (including Pronto) to detect sales
            # This enables sale detection, regular price capture, and percentage calculations for all stores
            api_result = self.scrape_freshop_store(url, store_type, product_name, driver)
            
            response_time = time.time() - start_time
            
            # Build result
            result = {
                'upc_plu': product_info.get('UPC/PLU', ''),
                'brand': product_info.get('Brand', ''),
                'product_name': product_name,
                'url': url,
                'store_type': store_type,
                'response_time': response_time,
                'scraped_at': datetime.now().isoformat(),
                'expected_price': product_info.get('Reg Retail'),
                'method': api_result.get('method', 'unknown')
            }
            
            if 'error' in api_result:
                result.update({
                    'status': 'failed',
                    'error': api_result['error'],
                    'price_found': None,
                    'price_match': False
                })
                self.performance_stats['failed_requests'] += 1
            else:
                result.update({
                    'status': 'success',
                    'price_found': api_result.get('price_found'),
                    'error': None
                })
                
                # Preserve comprehensive pricing data from scraping
                comprehensive_fields = ['sale_price', 'regular_price', 'original_price', 
                                       'sale_percentage', 'sale_start_date', 'sale_end_date', 'is_on_sale', 'website_sku']
                for field in comprehensive_fields:
                    if field in api_result:
                        result[field] = api_result[field]
                
                # Calculate price match
                expected_price = result['expected_price']
                found_price = result['price_found']
                
                if found_price and expected_price and isinstance(expected_price, (int, float)):
                    tolerance = 0.05  # 5% tolerance
                    price_diff = abs(found_price - expected_price) / expected_price
                    result['price_match'] = price_diff <= tolerance
                    result['price_difference'] = found_price - expected_price
                else:
                    result['price_match'] = False
                    result['price_difference'] = None
            
            return result
            
        except Exception as e:
            response_time = time.time() - start_time
            self.logger.error(f"Error processing {url}: {e}")
            self.performance_stats['failed_requests'] += 1
            
            return {
                'upc_plu': product_info.get('UPC/PLU', ''),
                'brand': product_info.get('Brand', ''),
                'product_name': product_name,
                'url': url,
                'store_type': store_type,
                'status': 'error',
                'error': str(e),
                'price_found': None,
                'expected_price': product_info.get('Reg Retail'),
                'price_match': False,
                'response_time': response_time,
                'method': 'error',
                'scraped_at': datetime.now().isoformat()
            }
    
    def monitor_products(self):
        """Monitor all products in the Excel file with persistent Chrome for MarketPlace"""
        try:
            df = pd.read_excel(self.input_file)
            self.logger.info(f"Loaded {len(df)} products from {self.input_file}")
            
            start_time = time.time()
            # FOCUS ON 3 RELIABLE STORES - Skip MarketPlace and Miles
            site_columns = ['HH', 'Drop It', 'Pronto']
            
            self.logger.info("🎯 FOCUSING ON 3 STORES TONIGHT - HH, Drop It, and Pronto (Skipping MarketPlace and Miles)")
            
            # No MarketPlace processing tonight
            
            # Process other stores (non-MarketPlace) - existing logic
            for index, row in df.iterrows():
                product_info = row.to_dict()
                product_name = f"{product_info.get('Brand', '')} {product_info.get('Long Description', '')}".strip()
                
                self.logger.info(f"\n📦 Processing Product {index + 1}: {product_name}")
                
                # Process reliable stores (skip MarketPlace and Miles)
                for site_col in ['HH', 'Drop It', 'Pronto']:
                    url = product_info.get(site_col)
                    
                    # Clean store type mapping
                    store_type_mapping = {
                        'HH': 'hh',
                        'Drop It': 'dropit', 
                        'Miles': 'miles',
                        'Pronto': 'pronto'
                    }
                    store_type = store_type_mapping.get(site_col, site_col.lower())
                    
                    self.logger.info(f"  🏪 Store: {site_col}")
                    
                    if not url or pd.isna(url):
                        self.logger.warning(f"    ❌ Failed: No URL provided")
                        continue
                    
                    result = self.process_single_url(url, store_type, product_info, product_name)
                    self.add_result_with_autosave(result)
                    
                    # Log result
                    if result['status'] == 'success':
                        method = result.get('method', 'unknown')
                        self.logger.info(f"    ✅ Success via {method}: ${result['price_found']}")
                    else:
                        self.logger.warning(f"    ❌ Failed: {result.get('error', 'Unknown error')}")
                    
                    # Small delay between stores
                    time.sleep(random.uniform(1, 3))
            
            total_time = time.time() - start_time
            self.logger.info(f"\n📊 Monitoring completed in {total_time:.2f} seconds")
            
        except Exception as e:
            self.logger.error(f"Error in monitoring: {e}")
            raise
    
    def generate_report(self):
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Hybrid Monitor Report"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Headers with enhanced pricing information
        headers = [
            "upc_plu", "website_sku", "brand", "product_name", "store_type", "expected_price", 
            "price_found", "sale_price", "regular_price", "price_difference", 
            "expected_vs_found_percentage", "sale_percentage", "price_match", 
            "is_on_sale", "sale_start_date", "sale_end_date", "method", 
            "status", "error", "response_time", "scraped_at", "url"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows with enhanced pricing information
        for row_idx, result in enumerate(self.results, 2):
            # Calculate expected vs found percentage
            expected_vs_found_percentage = None
            expected_price = result.get('expected_price')
            price_found = result.get('price_found')
            
            if expected_price and price_found:
                try:
                    expected = float(expected_price)
                    found = float(price_found)
                    if expected > 0:
                        expected_vs_found_percentage = round(((found - expected) / expected) * 100, 1)
                except (ValueError, TypeError):
                    pass
            
            ws.cell(row=row_idx, column=1, value=result.get('upc_plu', ''))                    # upc_plu
            ws.cell(row=row_idx, column=2, value=result.get('website_sku', ''))               # website_sku
            ws.cell(row=row_idx, column=3, value=result.get('brand', ''))                     # brand
            ws.cell(row=row_idx, column=4, value=result.get('product_name', ''))              # product_name
            ws.cell(row=row_idx, column=5, value=result.get('store_type', '').lower())        # store_type
            ws.cell(row=row_idx, column=6, value=result.get('expected_price', ''))            # expected_price
            ws.cell(row=row_idx, column=7, value=result.get('price_found', ''))               # price_found
            ws.cell(row=row_idx, column=8, value=result.get('sale_price', ''))                # sale_price
            ws.cell(row=row_idx, column=9, value=result.get('regular_price', ''))             # regular_price
            ws.cell(row=row_idx, column=10, value=result.get('price_difference', ''))          # price_difference
            ws.cell(row=row_idx, column=11, value=expected_vs_found_percentage)               # expected_vs_found_percentage
            ws.cell(row=row_idx, column=12, value=result.get('sale_percentage', ''))          # sale_percentage
            ws.cell(row=row_idx, column=13, value=result.get('price_match', ''))              # price_match
            ws.cell(row=row_idx, column=14, value=result.get('is_on_sale', ''))               # is_on_sale
            ws.cell(row=row_idx, column=15, value=result.get('sale_start_date', ''))          # sale_start_date
            ws.cell(row=row_idx, column=16, value=result.get('sale_end_date', ''))            # sale_end_date
            ws.cell(row=row_idx, column=17, value=result.get('method', ''))                   # method
            ws.cell(row=row_idx, column=18, value=result.get('status', ''))                   # status
            ws.cell(row=row_idx, column=19, value=result.get('error', ''))                    # error
            ws.cell(row=row_idx, column=20, value=f"{result.get('response_time', 0):.2f}s")   # response_time
            ws.cell(row=row_idx, column=21, value=result.get('scraped_at', ''))               # scraped_at
            ws.cell(row=row_idx, column=22, value=result.get('url', ''))                      # url
            
            # Color coding for all 22 columns
            if result.get('status') == 'success':
                for col in range(1, 23):  # Updated to 22 columns
                    ws.cell(row=row_idx, column=col).fill = success_fill
            else:
                for col in range(1, 23):  # Updated to 22 columns
                    ws.cell(row=row_idx, column=col).fill = error_fill
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        successful = len([r for r in self.results if r['status'] == 'success'])
        
        summary_data = [
            ["Hybrid Monitoring Summary", ""],
            ["Report Generated", datetime.now().isoformat()],
            ["", ""],
            ["Total Requests", self.performance_stats['total_requests']],
            ["Successful Results", successful],
            ["Success Rate", f"{(successful/len(self.results))*100:.1f}%" if self.results else "0%"],
            ["", ""],
            ["Method Breakdown", ""],
            ["Successful Scraping", self.performance_stats['successful_scrapes']],
            ["Successful API", self.performance_stats['successful_apis']],
            ["Successful OCR", self.performance_stats['successful_ocr']],
            ["Failed Requests", self.performance_stats['failed_requests']],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_idx, column=2, value=value)
        
        wb.save(self.output_file)
        self.logger.info(f"Report saved to {self.output_file}")
    
    def run(self):
        """Run the complete hybrid monitoring process"""
        self.logger.info("🚀 Starting Hybrid E-commerce Monitoring")
        self.logger.info("📝 Method: Web scraping (20s delays) for Freshop + API for Pronto + OCR fallback")
        
        self.monitor_products()
        self.generate_report()
        
        # Final stats
        successful = len([r for r in self.results if r['status'] == 'success'])
        total = len(self.results)
        
        self.logger.info(f"\n📊 FINAL RESULTS:")
        self.logger.info(f"   Total URLs: {total}")
        self.logger.info(f"   Successful: {successful}")
        self.logger.info(f"   Success Rate: {(successful/total)*100:.1f}%")
        self.logger.info(f"   Scraping: {self.performance_stats['successful_scrapes']}")
        self.logger.info(f"   API: {self.performance_stats['successful_apis']}")
        self.logger.info(f"   OCR: {self.performance_stats['successful_ocr']}")
        self.logger.info(f"   Failed: {self.performance_stats['failed_requests']}")

def main():
    input_file = "/Users/pato/Downloads/Download June_Top_100_with_URLs_Accurate.xlsx"
    output_file = "/Users/pato/Full_Monitoring_Results.xlsx"
    
    monitor = HybridEcommerceMonitor(input_file, output_file)
    monitor.run()

if __name__ == "__main__":
    main()