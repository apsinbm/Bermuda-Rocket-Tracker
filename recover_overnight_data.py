#!/usr/bin/env python3
"""
Recover overnight monitoring data from logs and create comprehensive Excel report
"""
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def recover_overnight_data():
    """Extract all overnight data from logs and create Excel report"""
    
    print('🔍 Recovering overnight monitoring data from logs...')
    
    # Read the log file
    with open('hybrid_monitor.log', 'r') as f:
        log_content = f.read()
    
    # Extract all July 2nd data with better parsing
    results = []
    current_product = None
    current_store = None
    
    lines = log_content.split('\n')
    for i, line in enumerate(lines):
        if '2025-07-02' not in line:
            continue
        
        # Extract product processing
        if 'Processing Product' in line:
            match = re.search(r'Processing Product (\d+): (.+)', line)
            if match:
                current_product = match.group(2).strip()
        
        # Extract store being processed
        if 'Store:' in line and '🏪' in line:
            match = re.search(r'🏪 Store: (.+)', line)
            if match:
                current_store = match.group(1).strip()
        
        # Extract MarketPlace products
        if 'MP Product' in line:
            match = re.search(r'MP Product \d+/\d+: (.+)', line)
            if match:
                current_product = match.group(1).strip()
                current_store = 'MarketPlace'
        
        # Extract all data points
        timestamp_match = re.search(r'2025-07-02 (\d{2}:\d{2}:\d{2})', line)
        if not timestamp_match:
            continue
            
        timestamp = timestamp_match.group(1)
        
        # SKU extraction
        sku_match = re.search(r'✓.*SKU found: (\d+)', line)
        if sku_match:
            sku = sku_match.group(1)
            results.append({
                'timestamp': timestamp,
                'product_name': current_product or 'Unknown',
                'store_type': current_store or 'MarketPlace',
                'website_sku': sku,
                'data_type': 'sku'
            })
        
        # Success extraction
        success_match = re.search(r'✅.*Success.*: \$(\d+\.?\d*)', line)
        if success_match:
            price = float(success_match.group(1))
            method = 'scraping'
            
            if 'OCR' in line or 'ocr' in line:
                method = 'ocr'
            elif 'comprehensive_scraping' in line:
                method = 'comprehensive_scraping'
            elif 'MP Success' in line:
                current_store = 'MarketPlace'
                method = 'marketplace_scraping'
            
            results.append({
                'timestamp': timestamp,
                'product_name': current_product or 'Unknown',
                'store_type': current_store or 'MarketPlace',
                'price_found': price,
                'method': method,
                'data_type': 'price'
            })
        
        # Failed attempts
        if '❌' in line and 'Failed:' in line:
            error_match = re.search(r'❌.*Failed: (.+)', line)
            if error_match:
                error = error_match.group(1).strip()
                results.append({
                    'timestamp': timestamp,
                    'product_name': current_product or 'Unknown',
                    'store_type': current_store or 'Unknown',
                    'error': error,
                    'data_type': 'error'
                })
    
    print(f'✅ Extracted {len(results)} raw data points')
    
    # Combine data points by product and store
    combined_data = {}
    
    for result in results:
        # Create unique key for each product-store combination
        key = f"{result['product_name']}_{result['store_type']}_{result['timestamp'][:5]}"
        
        if key not in combined_data:
            combined_data[key] = {
                'upc_plu': '',
                'website_sku': '',
                'brand': '',
                'product_name': result['product_name'],
                'store_type': result['store_type'].lower(),
                'expected_price': '',
                'price_found': '',
                'sale_price': '',
                'regular_price': '',
                'price_difference': '',
                'expected_vs_found_percentage': '',
                'sale_percentage': '',
                'price_match': '',
                'is_on_sale': '',
                'sale_start_date': '',
                'sale_end_date': '',
                'method': '',
                'status': 'failed',
                'error': '',
                'response_time': '',
                'scraped_at': f"2025-07-02 {result['timestamp']}",
                'url': ''
            }
        
        # Update with specific data
        if result['data_type'] == 'sku':
            combined_data[key]['website_sku'] = result.get('website_sku', '')
        elif result['data_type'] == 'price':
            combined_data[key]['price_found'] = result.get('price_found', '')
            combined_data[key]['method'] = result.get('method', '')
            combined_data[key]['status'] = 'success'
        elif result['data_type'] == 'error':
            combined_data[key]['error'] = result.get('error', '')
            combined_data[key]['status'] = 'failed'
    
    # Convert to list and clean up
    final_results = []
    for data in combined_data.values():
        # Extract brand and clean product name
        product_parts = data['product_name'].split()
        if len(product_parts) > 1:
            data['brand'] = product_parts[0]
            data['product_name'] = ' '.join(product_parts[1:]) if len(product_parts) > 1 else product_parts[0]
        
        # Set price match if we have a price
        if data['price_found']:
            data['price_match'] = 'Found'
        else:
            data['price_match'] = 'Not Found'
        
        final_results.append(data)
    
    print(f'✅ Created {len(final_results)} final records')
    
    # Create comprehensive Excel report
    df = pd.DataFrame(final_results)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Overnight Monitoring Results"
    
    # Headers
    headers = [
        "upc_plu", "website_sku", "brand", "product_name", "store_type", "expected_price", 
        "price_found", "sale_price", "regular_price", "price_difference", 
        "expected_vs_found_percentage", "sale_percentage", "price_match", 
        "is_on_sale", "sale_start_date", "sale_end_date", "method", 
        "status", "error", "response_time", "scraped_at", "url"
    ]
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, result in enumerate(final_results, 2):
        ws.cell(row=row_idx, column=1, value=result.get('upc_plu', ''))
        ws.cell(row=row_idx, column=2, value=result.get('website_sku', ''))
        ws.cell(row=row_idx, column=3, value=result.get('brand', ''))
        ws.cell(row=row_idx, column=4, value=result.get('product_name', ''))
        ws.cell(row=row_idx, column=5, value=result.get('store_type', ''))
        ws.cell(row=row_idx, column=6, value=result.get('expected_price', ''))
        ws.cell(row=row_idx, column=7, value=result.get('price_found', ''))
        ws.cell(row=row_idx, column=8, value=result.get('sale_price', ''))
        ws.cell(row=row_idx, column=9, value=result.get('regular_price', ''))
        ws.cell(row=row_idx, column=10, value=result.get('price_difference', ''))
        ws.cell(row=row_idx, column=11, value=result.get('expected_vs_found_percentage', ''))
        ws.cell(row=row_idx, column=12, value=result.get('sale_percentage', ''))
        ws.cell(row=row_idx, column=13, value=result.get('price_match', ''))
        ws.cell(row=row_idx, column=14, value=result.get('is_on_sale', ''))
        ws.cell(row=row_idx, column=15, value=result.get('sale_start_date', ''))
        ws.cell(row=row_idx, column=16, value=result.get('sale_end_date', ''))
        ws.cell(row=row_idx, column=17, value=result.get('method', ''))
        ws.cell(row=row_idx, column=18, value=result.get('status', ''))
        ws.cell(row=row_idx, column=19, value=result.get('error', ''))
        ws.cell(row=row_idx, column=20, value=result.get('response_time', ''))
        ws.cell(row=row_idx, column=21, value=result.get('scraped_at', ''))
        ws.cell(row=row_idx, column=22, value=result.get('url', ''))
        
        # Color coding
        if result.get('status') == 'success':
            for col in range(1, 23):
                ws.cell(row=row_idx, column=col).fill = success_fill
        else:
            for col in range(1, 23):
                ws.cell(row=row_idx, column=col).fill = error_fill
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    filename = 'Recovered_Overnight_Results.xlsx'
    wb.save(filename)
    
    print(f'\\n✅ COMPREHENSIVE EXCEL REPORT CREATED: {filename}')
    
    # Generate summary statistics
    successful = [r for r in final_results if r['status'] == 'success']
    with_sku = [r for r in final_results if r['website_sku']]
    stores = {}
    
    for result in final_results:
        store = result['store_type']
        if store not in stores:
            stores[store] = {'total': 0, 'success': 0}
        stores[store]['total'] += 1
        if result['status'] == 'success':
            stores[store]['success'] += 1
    
    print(f'\\n📊 COMPREHENSIVE ANALYSIS:')
    print(f'   Total Records: {len(final_results)}')
    print(f'   Successful Price Captures: {len(successful)}')
    print(f'   Records with Website SKUs: {len(with_sku)}')
    print(f'   Overall Success Rate: {len(successful)/len(final_results)*100:.1f}%')
    print(f'\\n🏪 STORE BREAKDOWN:')
    for store, stats in stores.items():
        success_rate = stats['success']/stats['total']*100 if stats['total'] > 0 else 0
        print(f'   {store.capitalize()}: {stats["success"]}/{stats["total"]} ({success_rate:.1f}%)')
    
    print(f'\\n💰 SAMPLE SUCCESSFUL CAPTURES:')
    for i, result in enumerate([r for r in final_results if r['status'] == 'success'][:10]):
        sku = result['website_sku'] if result['website_sku'] else 'No SKU'
        print(f'   {i+1}. {result["product_name"]} ({result["store_type"]}): ${result["price_found"]} - SKU: {sku}')
    
    return final_results

if __name__ == "__main__":
    recover_overnight_data()